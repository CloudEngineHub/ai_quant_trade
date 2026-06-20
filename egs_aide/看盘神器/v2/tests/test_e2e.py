# -*- coding: utf-8 -*-
"""端到端测试：验证完整工作流

测试流程:
    1. 从 YAML 加载配置
    2. 生成 Excel 模板
    3. 验证模板 Sheet 结构
    4. 用 mock 数据模拟完整刷新流程（4个 Sheet Handler）
    5. 验证数据正确写入

运行: python tests/test_e2e.py
"""
import os
import sys
import tempfile
import shutil
from unittest.mock import MagicMock

# 将 src 加入路径
_SRC_DIR = os.path.join(os.path.dirname(__file__), "..", "src")
_SRC_DIR = os.path.abspath(_SRC_DIR)
if _SRC_DIR not in sys.path:
    sys.path.insert(0, _SRC_DIR)

import pandas as pd
from openpyxl import load_workbook

from excel_monitor.config_loader import AppConfig, load_config
from excel_monitor.core.data_provider import DataProvider
from excel_monitor.core.alert_checker import AlertChecker, AlertCondition
from excel_monitor.sheets.market_overview import MarketOverviewSheet
from excel_monitor.sheets.detailed_quotes import DetailedQuotesSheet
from excel_monitor.sheets.news_sheet import NewsSheet
from excel_monitor.sheets.custom_watch import CustomWatchSheet
from excel_monitor.utils.template_generator import create_template


# ===== 测试工具 =====

def _section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def _assert(condition, msg):
    if condition:
        print(f"  [PASS] {msg}")
    else:
        print(f"  [FAIL] {msg}")
        raise AssertionError(msg)


# ===== 测试用例 =====

def test_e2e_config_loading():
    """E2E-1: 从 config.yaml 加载配置"""
    _section("E2E-1: YAML 配置加载")
    yaml_path = os.path.join(
        os.path.dirname(__file__), "..", "config.yaml"
    )
    yaml_path = os.path.abspath(yaml_path)

    cfg = load_config(yaml_path)

    _assert(cfg.refresh_interval == 3, f"刷新间隔 = 3 (实际: {cfg.refresh_interval})")
    _assert(cfg.excel_template == "看盘模板.xlsx", f"模板名 = 看盘模板.xlsx")
    _assert(len(cfg.market_indices) == 7, f"指数数量 = 7 (实际: {len(cfg.market_indices)})")
    _assert("上证指数" in cfg.market_indices, "包含上证指数")
    _assert(len(cfg.watch_stocks) == 6, f"自选股数量 = 6 (实际: {len(cfg.watch_stocks)})")
    _assert(len(cfg.custom_watch_columns) == 13, f"定制列数 = 13 (实际: {len(cfg.custom_watch_columns)})")
    _assert(cfg.sheets["market_overview"] == "大盘", "大盘 Sheet 名称正确")
    _assert(cfg.sheets["custom_watch"] == "个性定制看盘", "定制看盘 Sheet 名称正确")
    print("  --> 配置加载通过")


def test_e2e_template_generation():
    """E2E-2: 生成 Excel 模板并验证结构"""
    _section("E2E-3: Excel 模板生成与结构验证")

    tmp_dir = tempfile.mkdtemp(prefix="e2e_test_")
    try:
        template_path = os.path.join(tmp_dir, "看盘模板.xlsx")
        cfg = load_config(os.path.join(
            os.path.dirname(__file__), "..", "config.yaml"
        ))

        # 生成模板
        create_template(template_path, cfg)
        _assert(os.path.exists(template_path), "模板文件已生成")

        # 用 openpyxl 验证结构
        wb = load_workbook(template_path)
        sheet_names = wb.sheetnames
        _assert(len(sheet_names) == 5, f"Sheet 数量 = 5 (实际: {len(sheet_names)})")
        _assert("大盘" in sheet_names, "包含'大盘' Sheet")
        _assert("详细行情" in sheet_names, "包含'详细行情' Sheet")
        _assert("新闻" in sheet_names, "包含'新闻' Sheet")
        _assert("个性定制看盘" in sheet_names, "包含'个性定制看盘' Sheet")
        _assert("配置" in sheet_names, "包含'配置' Sheet")

        # 验证"详细行情" Sheet 有表头和示例数据
        ws_detail = wb["详细行情"]
        _assert(ws_detail["A1"].value == "代码", "详细行情 A1 = '代码'")
        _assert(ws_detail["B1"].value == "名称", "详细行情 B1 = '名称'")
        _assert(ws_detail["A2"].value == "中国平安", "详细行情 A2 = '中国平安'")

        # 验证"个性定制看盘" Sheet 表头
        ws_custom = wb["个性定制看盘"]
        _assert(ws_custom["A1"].value == "代码", "定制看盘 A1 = '代码'")
        _assert(ws_custom["B1"].value == "名称", "定制看盘 B1 = '名称'")
        _assert(ws_custom["M1"].value == "刷新时间", "定制看盘 M1 = '刷新时间'")
        # 验证预警条件列表头（N-Q 列，橙色填充）
        _assert(ws_custom["N1"].value == "涨跌幅下限", "定制看盘 N1 = '涨跌幅下限'")
        _assert(ws_custom["O1"].value == "涨跌幅上限", "定制看盘 O1 = '涨跌幅上限'")
        _assert(ws_custom["P1"].value == "价格下限", "定制看盘 P1 = '价格下限'")
        _assert(ws_custom["Q1"].value == "价格上限", "定制看盘 Q1 = '价格上限'")
        # 验证预警列表头为橙色
        alert_fill = ws_custom["N1"].fill
        _assert(alert_fill.start_color.rgb in ("FFED7D31", "00ED7D31"),
                f"预警列表头填充色 = ED7D31 (实际: {alert_fill.start_color.rgb})")
        # 验证示例预警条件
        _assert(ws_custom["O2"].value == 5.0, "中国平安 涨跌幅上限 = 5.0")
        _assert(ws_custom["P2"].value == 40.0, "中国平安 价格下限 = 40.0")
        _assert(ws_custom["Q3"].value == 2000.0, "贵州茅台 价格上限 = 2000.0")

        # 验证"配置" Sheet
        ws_config = wb["配置"]
        _assert(ws_config["A2"].value == "配置项", "配置 A2 = '配置项'")
        _assert(ws_config["B2"].value == "值", "配置 B2 = '值'")
        _assert(ws_config["A3"].value == "刷新间隔(秒)", "配置 A3 = '刷新间隔(秒)'")
        _assert(ws_config["B3"].value == 3, "配置 B3 = 3")
        _assert(ws_config["A4"].value == "预警弹窗", "配置 A4 = '预警弹窗'")
        _assert(ws_config["A6"].value == "配置重载间隔", "配置 A6 = '配置重载间隔'")
        _assert(ws_config["A9"].value == "自选股", "配置 A9 = '自选股'")
        _assert(ws_config["B9"].value == "指数", "配置 B9 = '指数'")
        _assert(ws_config["A10"].value == "中国平安", "配置 A10 = '中国平安'")
        _assert(ws_config["B10"].value == "上证指数", "配置 B10 = '上证指数'")

        # 验证表头样式（蓝色填充）
        fill = ws_detail["A1"].fill
        _assert(fill.start_color.rgb in ("FF4472C4", "004472C4"),
                f"表头填充色 = 4472C4 (实际: {fill.start_color.rgb})")

        print("  --> 模板生成与结构验证通过")
    finally:
        shutil.rmtree(tmp_dir)


def test_e2e_market_overview_full_refresh():
    """E2E-3: 大盘 Sheet 完整刷新流程"""
    _section("E2E-3: 大盘 Sheet 完整刷新流程")

    cfg = AppConfig()
    mock_excel = MagicMock()
    mock_data = MagicMock()

    # 模拟 qstock 返回数据
    mock_data.get_index_realtime.return_value = pd.DataFrame({
        "名称": ["上证指数", "深证成指", "创业板指"],
        "最新": [3200.0, 10500.0, 2100.0],
        "涨幅": [0.52, -0.31, 1.23],
    })
    mock_data.get_industry_boards.return_value = pd.DataFrame({
        "名称": ["银行", "证券", "半导体"],
        "涨幅": [1.2, -0.5, 2.8],
        "最新": [5200, 1800, 9500],
    })
    mock_data.get_concept_boards.return_value = pd.DataFrame({
        "名称": ["锂电池", "光伏", "AI算力"],
        "涨幅": [3.1, -1.2, 4.5],
    })
    mock_data.get_limit_up_pool.return_value = pd.DataFrame({
        "代码": ["000001", "600519"],
        "名称": ["平安银行", "贵州茅台"],
        "涨跌幅": [10.0, 10.0],
    })

    sheet = MarketOverviewSheet("大盘", mock_excel, mock_data, cfg)
    sheet.sheet = MagicMock()

    sheet.refresh()

    # 验证所有数据源被调用
    _assert(mock_data.get_index_realtime.call_count == 1, "指数行情被调用")
    _assert(mock_data.get_industry_boards.call_count == 1, "行业板块被调用")
    _assert(mock_data.get_concept_boards.call_count == 1, "概念板块被调用")
    _assert(mock_data.get_limit_up_pool.call_count == 1, "涨停板被调用")

    # 验证写入被调用4次（指数+行业+概念+涨停，通过 excel_mgr.write_df）
    _assert(mock_excel.write_df.call_count == 4, f"Excel 写入 4次 (实际: {mock_excel.write_df.call_count})")

    print("  --> 大盘 Sheet 刷新流程通过")


def test_e2e_detailed_quotes_full_flow():
    """E2E-4: 详细行情 Sheet 完整流程（init + refresh）"""
    _section("E2E-4: 详细行情 Sheet 完整流程")

    cfg = AppConfig()
    mock_excel = MagicMock()
    mock_data = MagicMock()

    # init 阶段：从 Excel 读取自选股
    mock_excel.sheet_to_df.return_value = (
        pd.DataFrame({
            "代码": ["中国平安", "贵州茅台", "宁德时代"],
            "名称": ["中国平安", "贵州茅台", "宁德时代"],
        }),
        4, 2,
    )

    # refresh 阶段：实时行情
    mock_data.get_stock_realtime.return_value = pd.DataFrame({
        "代码": ["中国平安", "贵州茅台", "宁德时代"],
        "名称": ["中国平安", "贵州茅台", "宁德时代"],
        "最新": [45.6, 1820.0, 210.5],
        "涨幅": [1.5, 0.8, -0.3],
        "时间": ["10:30:00", "10:30:00", "10:30:00"],
    })
    mock_data.get_billboard.return_value = pd.DataFrame({
        "代码": ["000001"], "名称": ["平安银行"], "净买入": [5000000],
    })
    mock_data.get_realtime_change.return_value = pd.DataFrame({
        "代码": ["600519"], "名称": ["贵州茅台"], "异动": ["火箭发射"],
    })

    sheet = DetailedQuotesSheet("详细行情", mock_excel, mock_data, cfg)
    sheet.sheet = MagicMock()

    # 执行 init
    sheet.init()
    _assert(len(sheet._stock_codes) == 3, f"自选股加载 3 只 (实际: {len(sheet._stock_codes)})")
    _assert("中国平安" in sheet._stock_codes, "包含中国平安")

    # 执行 refresh
    sheet.refresh()

    _assert(mock_data.get_stock_realtime.call_count == 1, "个股行情被调用")
    _assert(mock_data.get_billboard.call_count == 1, "龙虎榜被调用")
    _assert(mock_data.get_realtime_change.call_count == 1, "盘口异动被调用")

    print("  --> 详细行情 Sheet 流程通过")


def test_e2e_news_full_refresh():
    """E2E-5: 新闻 Sheet 完整刷新流程"""
    _section("E2E-5: 新闻 Sheet 完整刷新流程")

    cfg = AppConfig()
    cfg.news_max_rows = 3  # 限制条数便于验证

    mock_excel = MagicMock()
    mock_data = MagicMock()

    # 模拟 5 条新闻（应被截断为 3 条）
    mock_data.get_news_cls.return_value = pd.DataFrame({
        "发布时间": ["10:30", "10:25", "10:20", "10:15", "10:10"],
        "标题": ["新闻1", "新闻2", "新闻3", "新闻4", "新闻5"],
    })
    mock_data.get_news_js.return_value = pd.DataFrame({
        "时间": ["10:28"],
        "内容": ["快讯内容"],
    })

    sheet = NewsSheet("新闻", mock_excel, mock_data, cfg)
    sheet.sheet = MagicMock()

    sheet.refresh()

    _assert(mock_data.get_news_cls.call_count == 1, "财联社电报被调用")
    _assert(mock_data.get_news_js.call_count == 1, "市场快讯被调用")

    # 验证新闻条数被限制（第一次调用是财联社，第二次是快讯）
    cls_df = mock_excel.write_df.call_args_list[0][0][1]
    _assert(len(cls_df) == 3, f"财联社新闻截断为 3 条 (实际: {len(cls_df)})")
    js_df = mock_excel.write_df.call_args_list[1][0][1]
    _assert(len(js_df) == 1, f"市场快讯 1 条 (实际: {len(js_df)})")

    print("  --> 新闻 Sheet 刷新流程通过")


def test_e2e_custom_watch_full_flow():
    """E2E-6: 个性定制看盘 Sheet 完整流程"""
    _section("E2E-6: 个性定制看盘 Sheet 完整流程")

    cfg = AppConfig()
    mock_excel = MagicMock()

    # 使用真实的 DataProvider（纯逻辑方法不触发网络）
    real_dp = DataProvider()
    mock_data = MagicMock(wraps=real_dp)

    mock_excel.sheet_to_df.return_value = (
        pd.DataFrame({
            "代码": ["中国平安", "贵州茅台"],
            "名称": ["中国平安", "贵州茅台"],
        }),
        3, 2,
    )

    # qstock 返回完整列
    mock_data.get_stock_realtime.return_value = pd.DataFrame({
        "代码": ["中国平安", "贵州茅台"],
        "名称": ["中国平安", "贵州茅台"],
        "最新": [45.0, 1800.0],
        "涨幅": [1.5, 0.8],
        "涨跌额": [0.67, 14.3],
        "成交量": [100000, 5000],
        "成交额": [4500000, 9000000],
        "换手率": [0.5, 0.04],
        "最高": [45.5, 1810.0],
        "最低": [44.2, 1790.0],
        "开盘": [44.5, 1795.0],
        "昨收": [44.33, 1785.7],
        "时间": ["10:30:00", "10:30:00"],
        "多余列": [0, 0],
    })

    sheet = CustomWatchSheet("个性定制看盘", mock_excel, mock_data, cfg)
    sheet.sheet = MagicMock()

    sheet.init()
    sheet.refresh()

    _assert(mock_data.get_stock_realtime.call_count == 1, "个股行情被调用")
    _assert(mock_data.filter_columns.call_count == 1, "列过滤被调用")

    # 验证写入的 DataFrame
    written_df = mock_excel.write_df.call_args[0][1]
    _assert("刷新时间" in written_df.columns, "包含'刷新时间'列")
    _assert("多余列" not in written_df.columns, "不包含'多余列'")
    _assert(len(written_df) == 2, f"写入 2 行数据 (实际: {len(written_df)})")

    print("  --> 个性定制看盘 Sheet 流程通过")


def test_e2e_alert_full_flow():
    """E2E-7: 预警完整流程（条件加载 → 检查 → 高亮 → 弹窗消息）"""
    _section("E2E-7: 预警完整流程")

    cfg = AppConfig()
    cfg.alert_popup_enabled = False  # 测试中禁用实际弹窗
    mock_excel = MagicMock()

    real_dp = DataProvider()
    mock_data = MagicMock(wraps=real_dp)

    # Excel 中有预警条件
    mock_excel.sheet_to_df.return_value = (
        pd.DataFrame({
            "代码": ["中国平安", "贵州茅台", "宁德时代"],
            "名称": ["中国平安", "贵州茅台", "宁德时代"],
            "涨跌幅下限": [None, None, -5.0],
            "涨跌幅上限": [3.0, None, None],
            "价格下限": [40.0, None, None],
            "价格上限": [None, 2000.0, None],
        }),
        4, 6,
    )

    # 实时行情：中国平安涨5%(超涨幅上限)、茅台1850(正常)、宁德-6%(低于涨幅下限)
    mock_data.get_stock_realtime.return_value = pd.DataFrame({
        "代码": ["中国平安", "贵州茅台", "宁德时代"],
        "名称": ["中国平安", "贵州茅台", "宁德时代"],
        "最新": [45.0, 1850.0, 210.0],
        "涨幅": [5.0, 0.8, -6.0],
        "时间": ["10:30:00", "10:30:00", "10:30:00"],
    })

    sheet = CustomWatchSheet("个性定制看盘", mock_excel, mock_data, cfg)
    sheet.sheet = MagicMock()

    # init: 加载预警条件
    sheet.init()
    _assert(len(sheet._alert_conditions) == 3, f"加载 3 条预警条件 (实际: {len(sheet._alert_conditions)})")

    # refresh: 刷新 + 检查预警
    sheet.refresh()

    # 验证高亮被调用（中国平安 + 宁德时代 = 2 只触发）
    _assert(mock_excel.highlight_row.call_count == 2,
            f"高亮 2 行 (实际: {mock_excel.highlight_row.call_count})")

    # 验证清除高亮被调用
    mock_excel.clear_highlight.assert_called_once()

    # 验证数据写入只清除数据列（不触碰预警条件列）
    clear_call = mock_excel.clear_range.call_args
    _assert(clear_call.kwargs.get("end_col") == 13,
            f"清除范围 end_col=13 (实际: {clear_call.kwargs.get('end_col')})")

    print("  --> 预警完整流程通过")


def test_e2e_main_entry_import():
    """E2E-8: main.py 入口可导入"""
    _section("E2E-8: main.py 入口导入验证")

    base_dir = os.path.join(os.path.dirname(__file__), "..")
    base_dir = os.path.abspath(base_dir)
    src_dir = os.path.join(base_dir, "src")

    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)

    # 验证所有模块可导入
    from excel_monitor.config_loader import AppConfig, load_config
    from excel_monitor.core.data_provider import DataProvider
    from excel_monitor.core.excel_manager import ExcelManager
    from excel_monitor.sheets.market_overview import MarketOverviewSheet
    from excel_monitor.sheets.detailed_quotes import DetailedQuotesSheet
    from excel_monitor.sheets.news_sheet import NewsSheet
    from excel_monitor.sheets.custom_watch import CustomWatchSheet
    from excel_monitor.utils.template_generator import create_template

    _assert(True, "所有模块导入成功")
    print("  --> 入口导入验证通过")


# ===== 主入口 =====

def run_all():
    print("\n" + "="*60)
    print("  Excel 盯盘工具 V2 - 端到端测试")
    print("="*60)

    tests = [
        test_e2e_config_loading,
        test_e2e_template_generation,
        test_e2e_market_overview_full_refresh,
        test_e2e_detailed_quotes_full_flow,
        test_e2e_news_full_refresh,
        test_e2e_custom_watch_full_flow,
        test_e2e_alert_full_flow,
        test_e2e_main_entry_import,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            passed += 1
        except Exception as e:
            failed += 1
            print(f"\n  [ERROR] {test.__name__}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*60}")
    print(f"  结果: {passed} 通过, {failed} 失败, 共 {len(tests)} 项")
    print(f"{'='*60}\n")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(run_all())
