# -*- coding: utf-8 -*-
"""生成 Excel 看盘模板（使用 openpyxl）

运行: python -m excel_monitor.utils.template_generator
或通过 main.py 自动调用
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from excel_monitor.config_loader import AppConfig


def _style_header(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4",
                            fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_template(output_path: str = None, config: AppConfig = None):
    """生成 Excel 模板

    Args:
        output_path: 输出路径，默认使用 config.excel_template
        config: 配置对象，默认使用 AppConfig()
    """
    cfg = config or AppConfig()
    path = output_path or os.path.join(os.getcwd(), cfg.excel_template)

    wb = Workbook()

    # === Sheet 1: 大盘 ===
    ws1 = wb.active
    ws1.title = cfg.sheets["market_overview"]
    ws1["A1"] = "=== 大盘总览（自动刷新）==="
    ws1["A1"].font = Font(bold=True, size=14)

    # === Sheet 2: 详细行情 ===
    ws2 = wb.create_sheet(cfg.sheets["detailed_quotes"])
    ws2["A1"] = "代码"
    ws2["B1"] = "名称"
    _style_header(ws2["A1"])
    _style_header(ws2["B1"])
    # 示例自选股
    ws2["A2"] = "中国平安"
    ws2["B2"] = "中国平安"
    ws2["A3"] = "贵州茅台"
    ws2["B3"] = "贵州茅台"

    # === Sheet 3: 新闻 ===
    ws3 = wb.create_sheet(cfg.sheets["news"])
    ws3["A1"] = "=== 财经新闻（自动刷新）==="
    ws3["A1"].font = Font(bold=True, size=14)

    # === Sheet 4: 个性定制看盘 ===
    ws4 = wb.create_sheet(cfg.sheets["custom_watch"])
    # 写入数据列表头
    headers = cfg.custom_watch_columns
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        _style_header(cell)
    # 写入预警条件列表头（紧挨数据列右侧，用橙色区分）
    alert_headers = cfg.alert_columns
    alert_start_col = len(headers) + 1
    for i, header in enumerate(alert_headers):
        cell = ws4.cell(row=1, column=alert_start_col + i, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31",
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 示例自选股
    ws4["A2"] = "中国平安"
    ws4["B2"] = "中国平安"
    # 示例预警条件：涨跌幅上限 5%，价格下限 40
    ws4.cell(row=2, column=alert_start_col + 1, value=5.0)   # 涨跌幅上限
    ws4.cell(row=2, column=alert_start_col + 2, value=40.0)  # 价格下限
    ws4["A3"] = "贵州茅台"
    ws4["B3"] = "贵州茅台"
    # 示例预警条件：价格上限 2000
    ws4.cell(row=3, column=alert_start_col + 3, value=2000.0)  # 价格上限

    # K 线图操作区域（第 20 行开始）
    ws4["A20"] = "=== K 线图操作区 ==="
    ws4["A20"].font = Font(bold=True, size=12, color="4472C4")
    ws4["A21"] = "选中股票行号:"
    ws4["B21"] = 2  # 默认第 2 行（第一只股票）
    ws4["A22"] = "提示: 修改 B21 为要画K线的股票行号，然后点击按钮"
    ws4["A22"].font = Font(size=9, color="888888")
    # 按钮占位（实际按钮由 xlwings 在运行时添加）
    ws4["D21"] = "[画K线按钮在程序启动后自动出现]"
    ws4["D21"].font = Font(size=9, color="888888", italic=True)
    # K 线图显示区域标题
    ws4["A25"] = "K 线图将显示在此区域下方 ↓"
    ws4["A25"].font = Font(size=10, color="888888")

    # === Sheet 5: 资金情绪 ===
    ws5 = wb.create_sheet(cfg.sheets["sentiment"])
    ws5["A1"] = "=== 资金 & 情绪（自动刷新，多数据源）==="
    ws5["A1"].font = Font(bold=True, size=14)
    ws5["A2"] = "数据来源: AKShare(北向资金/微博舆情/新闻情绪) + 东方财富(股吧热门)"
    ws5["A2"].font = Font(size=9, color="888888", italic=True)

    # === Sheet 6: 股票池 ===
    ws6 = wb.create_sheet(cfg.sheets["stock_pool"])
    ws6["A1"] = "搜索关键字:"
    ws6["A1"].font = Font(bold=True)
    ws6["B1"] = ""  # 用户输入区
    ws6["C1"] = "[搜索按钮在程序启动后自动出现]"
    ws6["C1"].font = Font(size=9, color="888888", italic=True)
    ws6["D1"] = "[显示全部按钮]"
    ws6["D1"].font = Font(size=9, color="888888", italic=True)
    ws6["A2"] = "提示: 在 B1 输入代码/名称/拼音首字母，点搜索按钮过滤"
    ws6["A2"].font = Font(size=9, color="888888", italic=True)
    # 表头（第 4 行）
    ws6["A4"] = "代码"
    ws6["B4"] = "名称"
    ws6["C4"] = "市场"
    _style_header(ws6["A4"])
    _style_header(ws6["B4"])
    _style_header(ws6["C4"])
    # 示例数据（程序启动后会被真实股票池覆盖）
    ws6["A5"] = "000001"
    ws6["B5"] = "平安银行"
    ws6["C5"] = "深圳"

    # === Sheet 7: 配置 ===
    ws7 = wb.create_sheet(cfg.sheets["config"])
    # 标题
    ws7["A1"] = "=== 全局配置 ==="
    ws7["A1"].font = Font(bold=True, size=14)
    # 键值对表头
    ws7["A2"] = "配置项"
    ws7["B2"] = "值"
    _style_header(ws7["A2"])
    _style_header(ws7["B2"])
    # 标量配置项
    ws7["A3"] = "刷新间隔(秒)"
    ws7["B3"] = cfg.refresh_interval
    ws7["A4"] = "预警弹窗"
    ws7["B4"] = "true" if cfg.alert_popup_enabled else "false"
    ws7["A5"] = "新闻条数"
    ws7["B5"] = cfg.news_max_rows
    ws7["A6"] = "配置重载间隔"
    ws7["B6"] = cfg.config_reload_interval
    ws7["A7"] = "情绪条数"
    ws7["B7"] = cfg.sentiment_max_rows
    ws7["A8"] = "资金情绪Sheet"
    ws7["B8"] = "true" if cfg.sentiment_sheet_enabled else "false"
    ws7["A9"] = "备选数据源"
    ws7["B9"] = ",".join(cfg.enabled_backup_sources)
    ws7["A10"] = "股票池Sheet"
    ws7["B10"] = "true" if cfg.stock_pool_sheet_enabled else "false"
    ws7["A11"] = "股票池缓存天数"
    ws7["B11"] = cfg.stock_pool_cache_days

    # 空行后是列表配置
    ws7["A13"] = "=== 列表配置 ==="
    ws7["A13"].font = Font(bold=True, size=14)
    ws7["A14"] = "自选股"
    ws7["B14"] = "指数"
    _style_header(ws7["A14"])
    _style_header(ws7["B14"])
    # 写入自选股和指数列表
    max_len = max(len(cfg.watch_stocks), len(cfg.market_indices))
    for i in range(max_len):
        row = 15 + i
        if i < len(cfg.watch_stocks):
            ws7.cell(row=row, column=1, value=cfg.watch_stocks[i])
        if i < len(cfg.market_indices):
            ws7.cell(row=row, column=2, value=cfg.market_indices[i])

    # 设置列宽
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

    ws7.column_dimensions["A"].width = 20
    ws7.column_dimensions["B"].width = 20

    wb.save(path)
    print(f"模板已生成: {path}")
    return path


if __name__ == "__main__":
    create_template()
